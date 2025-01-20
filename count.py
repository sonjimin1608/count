from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 웹 드라이버 경로 설정 (다운로드한 ChromeDriver 경로 입력)
driver_path = "/opt/homebrew/bin/chromedriver"
service = Service(driver_path)

# Chrome 옵션 설정 (헤드리스 모드)
options = webdriver.ChromeOptions()
options.add_argument("--headless")  # 헤드리스 모드 활성화
options.add_argument("--no-sandbox")  # 권한 문제 방지
options.add_argument("--disable-dev-shm-usage")  # 메모리 문제 방지
options.add_argument("--disable-gpu")  # GPU 비활성화 (Linux 환경에서 필요)

# 브라우저 열기 (헤드리스 모드 적용)
driver = webdriver.Chrome(service=service, options=options)
print()
current_date = datetime.now()
formatted_Date = current_date.strftime("%Y/%m/%d(%a)")
print(formatted_Date)
week = int(input("몇 주차 "))
print("---------------------------------------------------------------------------------------------------")

id = ["dk040106", "ryan040", "jamie1608", "sdi5595", "dydrms0417", "thdwogjs040923", "sevenrich03"]
file_path = "result.xlsx"
wb = load_workbook(file_path)
sheet = wb.active
# r = ["B", str(9 * (week - 1) + 2),":F", str(9 * (week - 1) + 8)]
# rng = "".join(r)
range_cells = sheet["B2:F8"] 
data = [[cell.value for cell in row] for row in range_cells]

for i in range(len(id)):

    # Solved.ac 프로필 페이지로 이동
    driver.get("https://solved.ac/profile/" + id[i])
    
    # 페이지 로드 대기
    wait = WebDriverWait(driver, 30)
    wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")        # 테이블 찾기 (테이블이 존재하는지 확인)
    table = driver.find_element(By.CLASS_NAME, "css-1cyj4c5")
    rows = table.find_elements(By.TAG_NAME, "tr")
    # print(f"ID: {id[i]} - 테이블 행 갯수: {len(rows)}")

    # 각 행의 데이터 추출
    gold = str(int(rows[3].find_element(By.CLASS_NAME, "css-1vpdhtk").text) - data[i][1]).rjust(3)
    platinum = str(int(rows[4].find_element(By.CLASS_NAME, "css-1vpdhtk").text) - data[i][2]).rjust(3)
    diamond = rows[5].find_element(By.CLASS_NAME, "css-1vpdhtk").text
    ruby = rows[6].find_element(By.CLASS_NAME, "css-1vpdhtk").text
    dia_above = str(int(diamond) + int(ruby) - data[i][3]).rjust(3)
    
    # 버튼 요소 대기 (CSS Selector 사용)
    button = wait.until(EC.element_to_be_clickable((By.XPATH,"//button[text()='자세히']")))
    
    # 버튼 클릭
    button.click()
    # print("버튼을 클릭했습니다.")
    time.sleep(1)  # 버튼 클릭 후 대기
    table = driver.find_element(By.CLASS_NAME, "css-1cyj4c5")
    rows = table.find_elements(By.TAG_NAME, "tr")
    # print(f"ID: {id[i]} - 테이블 행 갯수: {len(rows)}")

    # 각 행의 데이터 추출
    silver2 = str(int(rows[10].find_element(By.CLASS_NAME, "css-1vpdhtk").text)).rjust(3)
    silver1 = str(int(rows[11].find_element(By.CLASS_NAME, "css-1vpdhtk").text) - data[i][0]).rjust(3)
    silver1 = str(int(silver1) + int(silver2)).rjust(3)
    total = str(0.5 * int(silver1) + int(gold) + int(platinum) * 2 + int(dia_above) * 5).rjust(5)
    idid = id[i].ljust(15)
    
    print("|", idid, "| Silver :",silver1,
                     "| Gold :", gold,
                     "| Platinum :",platinum, 
                     "| Dia & Ruby :", dia_above,
                     "| Total :", total, "|")
    print("---------------------------------------------------------------------------------------------------")
    l = [int(silver1), int(gold), int(platinum), int(dia_above), float(total)]
    for j in range(5):
        cell = sheet.cell(row = 9 * week + 2 + i, column = 2 + j)
        cell.value = l[j]
    cell = sheet.cell(row = 9 * week + 2 + i, column = 6)
    if cell.value < 5:
        cell.fill = PatternFill(start_color = "d1162f", end_color = "d1162f", fill_type = "solid")
    else:
        cell.fill = PatternFill(start_color = "00962f", end_color = "00962f", fill_type = "solid")
print()
print()

file_path = "result.xlsx"
wb.save(file_path)
